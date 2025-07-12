"""
Export Utilities
Utilities for exporting data, charts, and generating reports.
"""

import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from pathlib import Path
import json
import logging
from typing import Dict, Any, List, Optional, Union
from datetime import datetime
import tempfile
import zipfile
import io

logger = logging.getLogger(__name__)

class ExportManager:
    """Manages various export operations."""
    
    def __init__(self, output_dir: str = "data/exports"):
        """
        Initialize export manager.
        
        Args:
            output_dir: Default output directory
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_dataframe_to_excel(self, 
                                dataframes: Dict[str, pd.DataFrame],
                                filename: str = None,
                                include_charts: bool = False) -> str:
        """
        Export multiple DataFrames to Excel with multiple sheets.
        
        Args:
            dataframes: Dictionary of sheet_name: DataFrame
            filename: Output filename (if None, auto-generated)
            include_charts: Whether to include chart sheets
            
        Returns:
            Path to exported file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"youtube_analytics_export_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                for sheet_name, df in dataframes.items():
                    # Clean sheet name (Excel has restrictions)
                    clean_sheet_name = sheet_name.replace('/', '_').replace('\\', '_')[:31]
                    df.to_excel(writer, sheet_name=clean_sheet_name, index=False)
                    
                    # Format the worksheet
                    worksheet = writer.sheets[clean_sheet_name]
                    self._format_excel_worksheet(worksheet, df)
            
            logger.info(f"Excel export completed: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def export_dataframe_to_csv(self, 
                               df: pd.DataFrame,
                               filename: str = None,
                               include_index: bool = False) -> str:
        """
        Export DataFrame to CSV.
        
        Args:
            df: DataFrame to export
            filename: Output filename
            include_index: Whether to include index
            
        Returns:
            Path to exported file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"youtube_data_export_{timestamp}.csv"
        
        filepath = self.output_dir / filename
        
        try:
            df.to_csv(filepath, index=include_index)
            logger.info(f"CSV export completed: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise
    
    def export_to_json(self, 
                      data: Dict[str, Any],
                      filename: str = None,
                      pretty_print: bool = True) -> str:
        """
        Export data to JSON.
        
        Args:
            data: Data to export
            filename: Output filename
            pretty_print: Whether to format JSON nicely
            
        Returns:
            Path to exported file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"youtube_analytics_results_{timestamp}.json"
        
        filepath = self.output_dir / filename
        
        try:
            with open(filepath, 'w') as f:
                if pretty_print:
                    json.dump(data, f, indent=2, default=str)
                else:
                    json.dump(data, f, default=str)
            
            logger.info(f"JSON export completed: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            raise
    
    def create_export_package(self, 
                            dataframes: Dict[str, pd.DataFrame],
                            charts: Dict[str, go.Figure],
                            analysis_results: Dict[str, Any],
                            package_name: str = None) -> str:
        """
        Create a complete export package with all data and charts.
        
        Args:
            dataframes: DataFrames to include
            charts: Charts to include
            analysis_results: Analysis results to include
            package_name: Name of the package
            
        Returns:
            Path to exported package
        """
        if package_name is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            package_name = f"youtube_analytics_package_{timestamp}.zip"
        
        package_path = self.output_dir / package_name
        
        try:
            with zipfile.ZipFile(package_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                # Add Excel file
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    for sheet_name, df in dataframes.items():
                        clean_name = sheet_name.replace('/', '_').replace('\\', '_')[:31]
                        df.to_excel(writer, sheet_name=clean_name, index=False)
                
                zf.writestr("data/youtube_analytics_data.xlsx", excel_buffer.getvalue())
                
                # Add charts
                chart_exporter = ChartExporter()
                for chart_name, chart_fig in charts.items():
                    if chart_fig is not None:
                        # HTML version
                        html_content = chart_exporter.export_to_html_string(chart_fig)
                        zf.writestr(f"charts/{chart_name}.html", html_content)
                        
                        # PNG version
                        try:
                            png_content = chart_exporter.export_to_png_bytes(chart_fig)
                            zf.writestr(f"charts/png/{chart_name}.png", png_content)
                        except Exception as e:
                            logger.warning(f"Could not export {chart_name} to PNG: {e}")
                
                # Add analysis results
                zf.writestr("analysis_results.json", 
                           json.dumps(analysis_results, indent=2, default=str))
                
                # Add README
                readme_content = self._generate_package_readme(dataframes, charts, analysis_results)
                zf.writestr("README.md", readme_content)
            
            logger.info(f"Export package created: {package_path}")
            return str(package_path)
            
        except Exception as e:
            logger.error(f"Error creating export package: {e}")
            raise
    
    def _format_excel_worksheet(self, worksheet, df: pd.DataFrame):
        """Format Excel worksheet for better presentation."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except ImportError:
            logger.warning("openpyxl styling not available, using basic formatting")
        except Exception as e:
            logger.warning(f"Error formatting Excel worksheet: {e}")
    
    def _generate_package_readme(self, 
                               dataframes: Dict[str, pd.DataFrame],
                               charts: Dict[str, go.Figure],
                               analysis_results: Dict[str, Any]) -> str:
        """Generate README for export package."""
        readme = f"""# YouTube Analytics Export Package

Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Contents

### Data Files
- `data/youtube_analytics_data.xlsx` - Complete dataset with multiple sheets

#### Data Sheets:
"""
        
        for sheet_name, df in dataframes.items():
            readme += f"- **{sheet_name}**: {len(df)} rows, {len(df.columns)} columns\n"
        
        readme += f"""
### Charts
- `charts/` - Interactive HTML charts
- `charts/png/` - Static PNG images

#### Available Charts:
"""
        
        for chart_name in charts.keys():
            if charts[chart_name] is not None:
                readme += f"- {chart_name}.html / {chart_name}.png\n"
        
        readme += f"""
### Analysis Results
- `analysis_results.json` - Complete analysis results in JSON format

## Usage

1. Open Excel file for data exploration
2. View HTML charts in web browser for interactive analysis
3. Use PNG charts for presentations
4. Load JSON file for programmatic analysis

## Analysis Summary
"""
        
        if 'summary_statistics' in analysis_results:
            summary = analysis_results['summary_statistics']
            if 'overview' in summary:
                overview = summary['overview']
                readme += f"""
### Key Metrics
- Total Videos: {overview.get('total_videos', 'N/A'):,}
- Total Views: {overview.get('total_views', 'N/A'):,}
- Total Likes: {overview.get('total_likes', 'N/A'):,}
- Total Comments: {overview.get('total_comments', 'N/A'):,}
"""
        
        readme += """
## Technical Notes

- Charts are created using Plotly for interactivity
- Data processed using pandas for accuracy
- All exports include data validation and quality checks

For questions or support, refer to the project documentation.
"""
        
        return readme


class ChartExporter:
    """Utilities for exporting charts in various formats."""
    
    def __init__(self):
        """Initialize chart exporter."""
        self.default_config = {
            'displayModeBar': False,
            'responsive': True
        }
    
    def export_to_html(self, 
                      fig: go.Figure, 
                      filepath: str,
                      include_plotlyjs: str = 'cdn',
                      config: Dict[str, Any] = None) -> str:
        """
        Export chart to HTML file.
        
        Args:
            fig: Plotly figure
            filepath: Output file path
            include_plotlyjs: How to include Plotly.js ('cdn', 'inline', etc.)
            config: Chart configuration
            
        Returns:
            Path to exported file
        """
        chart_config = {**self.default_config, **(config or {})}
        
        try:
            fig.write_html(
                filepath,
                include_plotlyjs=include_plotlyjs,
                config=chart_config
            )
            logger.info(f"Chart exported to HTML: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting chart to HTML: {e}")
            raise
    
    def export_to_html_string(self, 
                            fig: go.Figure,
                            include_plotlyjs: str = 'cdn',
                            config: Dict[str, Any] = None) -> str:
        """
        Export chart to HTML string.
        
        Args:
            fig: Plotly figure
            include_plotlyjs: How to include Plotly.js
            config: Chart configuration
            
        Returns:
            HTML string
        """
        chart_config = {**self.default_config, **(config or {})}
        
        try:
            return fig.to_html(
                include_plotlyjs=include_plotlyjs,
                config=chart_config
            )
            
        except Exception as e:
            logger.error(f"Error exporting chart to HTML string: {e}")
            raise
    
    def export_to_png(self, 
                     fig: go.Figure,
                     filepath: str,
                     width: int = 1200,
                     height: int = 800,
                     scale: float = 2.0) -> str:
        """
        Export chart to PNG file.
        
        Args:
            fig: Plotly figure
            filepath: Output file path
            width: Image width in pixels
            height: Image height in pixels
            scale: Scale factor for higher resolution
            
        Returns:
            Path to exported file
        """
        try:
            fig.write_image(
                filepath,
                format="png",
                width=width,
                height=height,
                scale=scale
            )
            logger.info(f"Chart exported to PNG: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting chart to PNG: {e}")
            logger.warning("Note: PNG export requires kaleido package: pip install kaleido")
            raise
    
    def export_to_png_bytes(self, 
                           fig: go.Figure,
                           width: int = 1200,
                           height: int = 800,
                           scale: float = 2.0) -> bytes:
        """
        Export chart to PNG bytes.
        
        Args:
            fig: Plotly figure
            width: Image width in pixels
            height: Image height in pixels
            scale: Scale factor for higher resolution
            
        Returns:
            PNG image as bytes
        """
        try:
            return fig.to_image(
                format="png",
                width=width,
                height=height,
                scale=scale
            )
            
        except Exception as e:
            logger.error(f"Error exporting chart to PNG bytes: {e}")
            raise
    
    def export_to_pdf(self, 
                     fig: go.Figure,
                     filepath: str,
                     width: int = 1200,
                     height: int = 800) -> str:
        """
        Export chart to PDF file.
        
        Args:
            fig: Plotly figure
            filepath: Output file path
            width: Image width in pixels
            height: Image height in pixels
            
        Returns:
            Path to exported file
        """
        try:
            fig.write_image(
                filepath,
                format="pdf",
                width=width,
                height=height
            )
            logger.info(f"Chart exported to PDF: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting chart to PDF: {e}")
            raise


class ReportGenerator:
    """Generates comprehensive reports from analytics data."""
    
    def __init__(self, template_dir: Optional[str] = None):
        """
        Initialize report generator.
        
        Args:
            template_dir: Directory containing report templates
        """
        self.template_dir = Path(template_dir) if template_dir else None
    
    def generate_summary_report(self, 
                              analysis_results: Dict[str, Any],
                              output_path: str = None) -> str:
        """
        Generate a comprehensive summary report.
        
        Args:
            analysis_results: Complete analysis results
            output_path: Output file path
            
        Returns:
            Path to generated report
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"data/exports/youtube_analytics_report_{timestamp}.html"
        
        try:
            report_html = self._create_html_report(analysis_results)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report_html)
            
            logger.info(f"Summary report generated: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating summary report: {e}")
            raise
    
    def generate_executive_summary(self, 
                                 analysis_results: Dict[str, Any]) -> str:
        """
        Generate an executive summary in markdown format.
        
        Args:
            analysis_results: Complete analysis results
            
        Returns:
            Executive summary as markdown string
        """
        try:
            summary = analysis_results.get('summary_statistics', {})
            
            markdown = f"""# YouTube Analytics Executive Summary

**Report Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Channel Overview
"""
            
            if 'overview' in summary:
                overview = summary['overview']
                markdown += f"""
- **Total Videos:** {overview.get('total_videos', 'N/A'):,}
- **Total Views:** {overview.get('total_views', 'N/A'):,}
- **Total Likes:** {overview.get('total_likes', 'N/A'):,}
- **Total Comments:** {overview.get('total_comments', 'N/A'):,}
- **Date Range:** {overview.get('date_range', {}).get('start', 'N/A')} to {overview.get('date_range', {}).get('end', 'N/A')}
"""
            
            if 'engagement_metrics' in summary:
                engagement = summary['engagement_metrics']
                markdown += f"""
## Engagement Metrics

- **Average Like Rate:** {engagement.get('average_like_rate', 0):.2f}%
- **Average Comment Rate:** {engagement.get('average_comment_rate', 0):.2f}%
- **Average Engagement Rate:** {engagement.get('average_engagement_rate', 0):.2f}%
- **Median Views:** {engagement.get('median_views', 0):,.0f}
"""
            
            if 'top_performers' in summary:
                top = summary['top_performers']
                markdown += f"""
## Top Performers

### Most Viewed Video
- **Title:** {top.get('most_viewed', {}).get('title', 'N/A')}
- **Views:** {top.get('most_viewed', {}).get('views', 0):,}

### Highest Like Rate
- **Title:** {top.get('highest_like_rate', {}).get('title', 'N/A')}
- **Rate:** {top.get('highest_like_rate', {}).get('rate', 0):.2f}%
"""
            
            # Add insights if available
            if 'insights' in analysis_results:
                insights = analysis_results['insights']
                markdown += "\n## Key Insights\n"
                
                for category, recommendations in insights.items():
                    if recommendations and category != 'error':
                        markdown += f"\n### {category.replace('_', ' ').title()}\n"
                        for rec in recommendations:
                            markdown += f"- {rec}\n"
            
            # Add ML results if available
            if 'ml_training' in analysis_results and 'error' not in analysis_results['ml_training']:
                ml_results = analysis_results['ml_training']
                if 'performance_metrics' in ml_results:
                    metrics = ml_results['performance_metrics']
                    markdown += f"""
## Machine Learning Model Performance

- **R² Score:** {metrics.get('r2_score', 0):.3f}
- **Mean Absolute Error:** {metrics.get('mae', 0):.0f} views
- **Root Mean Square Error:** {metrics.get('rmse', 0):.0f} views
"""
            
            return markdown
            
        except Exception as e:
            logger.error(f"Error generating executive summary: {e}")
            return f"Error generating executive summary: {str(e)}"
    
    def _create_html_report(self, analysis_results: Dict[str, Any]) -> str:
        """Create comprehensive HTML report."""
        
        html_template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>YouTube Analytics Report</title>
    <style>
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Roboto', sans-serif;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            line-height: 1.6;
            color: #333;
        }
        .header {
            text-align: center;
            margin-bottom: 40px;
            padding: 30px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-radius: 10px;
        }
        .header h1 {
            margin: 0;
            font-size: 2.5rem;
        }
        .header p {
            margin: 10px 0 0 0;
            opacity: 0.9;
        }
        .section {
            margin: 30px 0;
            padding: 20px;
            background: #f8f9fa;
            border-radius: 8px;
            border-left: 4px solid #007bff;
        }
        .section h2 {
            margin-top: 0;
            color: #007bff;
        }
        .metrics-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }
        .metric-card {
            background: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .metric-value {
            font-size: 2rem;
            font-weight: bold;
            color: #007bff;
            margin: 0;
        }
        .metric-label {
            font-size: 0.9rem;
            color: #6c757d;
            margin: 5px 0 0 0;
        }
        .insights-list {
            list-style: none;
            padding: 0;
        }
        .insights-list li {
            background: white;
            margin: 10px 0;
            padding: 15px;
            border-radius: 5px;
            border-left: 4px solid #28a745;
        }
        .table {
            width: 100%;
            border-collapse: collapse;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .table th,
        .table td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #dee2e6;
        }
        .table th {
            background-color: #007bff;
            color: white;
            font-weight: 600;
        }
        .footer {
            text-align: center;
            margin-top: 40px;
            padding: 20px;
            border-top: 1px solid #dee2e6;
            color: #6c757d;
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>📺 YouTube Analytics Report</h1>
        <p>Generated on {report_date}</p>
    </div>

    {content}

    <div class="footer">
        <p>Report generated by YouTube Analytics System</p>
    </div>
</body>
</html>
"""
        
        # Generate content sections
        content_sections = []
        
        # Overview section
        if 'summary_statistics' in analysis_results:
            content_sections.append(self._create_overview_section(analysis_results['summary_statistics']))
        
        # Insights section
        if 'insights' in analysis_results:
            content_sections.append(self._create_insights_section(analysis_results['insights']))
        
        # ML section
        if 'ml_training' in analysis_results and 'error' not in analysis_results['ml_training']:
            content_sections.append(self._create_ml_section(analysis_results['ml_training']))
        
        content = '\n'.join(content_sections)
        
        return html_template.format(
            report_date=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            content=content
        )
    
    def _create_overview_section(self, summary_stats: Dict[str, Any]) -> str:
        """Create overview section HTML."""
        html = '<div class="section"><h2>📊 Channel Overview</h2>'
        
        if 'overview' in summary_stats:
            overview = summary_stats['overview']
            html += '<div class="metrics-grid">'
            
            metrics = [
                ('Total Videos', f"{overview.get('total_videos', 0):,}"),
                ('Total Views', f"{overview.get('total_views', 0):,}"),
                ('Total Likes', f"{overview.get('total_likes', 0):,}"),
                ('Total Comments', f"{overview.get('total_comments', 0):,}")
            ]
            
            for label, value in metrics:
                html += f'''
                <div class="metric-card">
                    <p class="metric-value">{value}</p>
                    <p class="metric-label">{label}</p>
                </div>
                '''
            
            html += '</div>'
        
        html += '</div>'
        return html
    
    def _create_insights_section(self, insights: Dict[str, Any]) -> str:
        """Create insights section HTML."""
        html = '<div class="section"><h2>💡 Key Insights</h2>'
        
        for category, recommendations in insights.items():
            if recommendations and category != 'error':
                html += f'<h3>{category.replace("_", " ").title()}</h3>'
                html += '<ul class="insights-list">'
                
                for rec in recommendations:
                    html += f'<li>{rec}</li>'
                
                html += '</ul>'
        
        html += '</div>'
        return html
    
    def _create_ml_section(self, ml_results: Dict[str, Any]) -> str:
        """Create ML section HTML."""
        html = '<div class="section"><h2>🤖 Machine Learning Results</h2>'
        
        if 'performance_metrics' in ml_results:
            metrics = ml_results['performance_metrics']
            html += '<div class="metrics-grid">'
            
            ml_metrics = [
                ('R² Score', f"{metrics.get('r2_score', 0):.3f}"),
                ('MAE', f"{metrics.get('mae', 0):.0f}"),
                ('RMSE', f"{metrics.get('rmse', 0):.0f}"),
                ('Model Type', ml_results.get('model_type', 'N/A'))
            ]
            
            for label, value in ml_metrics:
                html += f'''
                <div class="metric-card">
                    <p class="metric-value">{value}</p>
                    <p class="metric-label">{label}</p>
                </div>
                '''
            
            html += '</div>'
        
        html += '</div>'
        return html
